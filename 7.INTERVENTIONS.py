import openpyxl
import os

def get_excel_filename(pattern):
    """Lire le nom du fichier Excel correspondant au pattern depuis excel_filenames.txt"""
    if os.path.exists('excel_filenames.txt'):
        with open('excel_filenames.txt', 'r') as f:
            filenames = [line.strip() for line in f.readlines()]
        
        # Chercher le fichier correspondant au pattern
        for filename in filenames:
            if pattern.lower() in filename.lower():
                return filename
    
    return None

# R2 Ouvrir le fichier SORTIE DU MOIS EN COURS.xlsx
sortie_workbook = openpyxl.load_workbook('SORTIE.xlsx')
sortie_sheet = sortie_workbook['INTERVENTIONS']

# R6 Ouvrir le fichier INTERVENTION ASTREINTE - Récupération automatique du nom
intervention_file_name = get_excel_filename('intervention astreinte')
if not intervention_file_name:
    print("Erreur : Impossible de trouver le fichier 'INTERVENTION ASTREINTE'")
    exit(1)

print(f"Ouverture du fichier : {intervention_file_name}")
intervention_workbook = openpyxl.load_workbook(intervention_file_name)
intervention_technique_sheet = intervention_workbook['TECHNIQUE']
intervention_administratif_sheet = intervention_workbook['ADMINISTRATIF']

# Créer un dictionnaire pour stocker les mois de référence
mois_ref = {}

# Créer un ensemble pour stocker les matricules du fichier ETAT DES HEURES SUPPLEMENTAIRES JANVIER 2025.xlsx
sortie_matricules = set()

# Parcourir les lignes de la feuille Interventions de ETAT DES HEURES SUPPLEMENTAIRES JANVIER 2025.xlsx
for row in sortie_sheet.iter_rows(min_row=2, values_only=True):
    matricule, mois_annee = row[:2]
    sortie_matricules.add(matricule)
    if matricule in mois_ref:
        mois_ref[matricule].append(mois_annee)
    else:
        mois_ref[matricule] = [mois_annee]

# Fonction pour mettre à jour les dates dans une feuille
def update_dates(sheet):
    for idx, row in enumerate(sheet.iter_rows(min_row=5, min_col=2, max_col=15, values_only=True), start=5):
        matricule = row[0]
        if matricule in sortie_matricules and matricule in mois_ref:
            mois_reference = ' | '.join([mois for mois in mois_ref[matricule]]).replace('|', ' | ')
            sheet.cell(row=idx, column=10, value=mois_reference)

# Mettre à jour les dates dans les feuilles TECHNIQUE et ADMINISTRATIF
update_dates(intervention_technique_sheet)
update_dates(intervention_administratif_sheet)

# R6 Enregistrer les modifications dans le fichier INTERVENTION ASTREINTE
intervention_workbook.save(intervention_file_name)
print(f"Fichier {intervention_file_name} mis à jour avec succès.")