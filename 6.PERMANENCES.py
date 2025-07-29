import openpyxl
import os

def get_excel_filename(pattern):
    """Lire le nom du fichier Excel correspondant au pattern depuis excel_filenames.txt"""
    if os.path.exists('excel_filenames.txt'):
        with open('excel_filenames.txt', 'r') as f:
            filenames = [line.strip() for line in f.readlines()]
        
        # Chercher le fichier correspondant au pattern
        for filename in filenames:
            if pattern.lower() in filename.lower():
                return filename
    
    return None

# R2 Ouvrir le fichier SORTIE DU MOIS EN COURS.xlsx
sortie_workbook = openpyxl.load_workbook('SORTIE.xlsx')
sortie_sheet = sortie_workbook['PERMANENCES']

# R5 Ouvrir le fichier INDEMNITE PERMANENCE - Récupération automatique du nom
indemnite_file_name = get_excel_filename('indemnite permanence')
if not indemnite_file_name:
    print("Erreur : Impossible de trouver le fichier 'INDEMNITE PERMANENCE'")
    exit(1)

print(f"Ouverture du fichier : {indemnite_file_name}")
indemnite_workbook = openpyxl.load_workbook(indemnite_file_name)
indemnite_technique_sheet = indemnite_workbook['TECHNIQUE']
indemnite_administratif_sheet = indemnite_workbook['ADMINISTRATIF']

# Créer un dictionnaire pour stocker les mois de référence
mois_ref = {}

# Créer un ensemble pour stocker les matricules du fichier ETAT DES HEURES SUPPLEMENTAIRES JANVIER 2025.xlsx
sortie_matricules = set()

# Parcourir les lignes de la feuille Permanences de ETAT DES HEURES SUPPLEMENTAIRES JANVIER 2025.xlsx
for row in sortie_sheet.iter_rows(min_row=2, values_only=True):
    matricule, mois_annee = row[:2]
    sortie_matricules.add(matricule)
    if matricule in mois_ref:
        mois_ref[matricule].append(mois_annee)
    else:
        mois_ref[matricule] = [mois_annee]

# Fonction pour mettre à jour les dates dans une feuille
def update_dates(sheet):
    for idx, row in enumerate(sheet.iter_rows(min_row=5, min_col=2, max_col=15, values_only=True), start=5):
        matricule = row[0]
        if matricule in sortie_matricules and matricule in mois_ref:
            mois_reference = ' | '.join([mois for mois in mois_ref[matricule]]).replace('|', ' | ')
            sheet.cell(row=idx, column=10, value=mois_reference)

# Mettre à jour les dates dans les feuilles TECHNIQUE et ADMINISTRATIF
update_dates(indemnite_technique_sheet)
update_dates(indemnite_administratif_sheet)

# R5 Enregistrer les modifications dans le fichier INDEMNITE PERMANENCE
indemnite_workbook.save(indemnite_file_name)
print(f"Fichier {indemnite_file_name} mis à jour avec succès.")