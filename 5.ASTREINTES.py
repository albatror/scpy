import openpyxl

# R2 Ouvrir le fichier SORTIE DU MOIS EN COURS.xlsx
sortie_workbook = openpyxl.load_workbook('SORTIE.xlsx')
sortie_sheet = sortie_workbook['ASTREINTES']

# R4 Ouvrir le fichier INDEMNITE ASTREINTE MOIS ANNEE
indemnite_workbook = openpyxl.load_workbook('INDEMNITE ASTREINTE SIEGE JUILLET 2025.xlsx')
indemnite_technique_sheet = indemnite_workbook['TECHNIQUE']
indemnite_administratif_sheet = indemnite_workbook['ADMINISTRATIF']

# Créer un dictionnaire pour stocker les mois de référence
mois_ref = {}

# Créer un ensemble pour stocker les matricules du fichier ETAT DES HEURES SUPPLEMENTAIRES JANVIER 2025.xlsx
sortie_matricules = set()

# Parcourir les lignes de la feuille Astreintes de ETAT DES HEURES SUPPLEMENTAIRES JANVIER 2025.xlsx
for row in sortie_sheet.iter_rows(min_row=2, values_only=True):
    matricule, mois_annee = row[:2]
    sortie_matricules.add(matricule)
    if matricule in mois_ref:
        mois_ref[matricule].append(mois_annee)
    else:
        mois_ref[matricule] = [mois_annee]

# Fonction pour mettre à jour les dates dans une feuille
def update_dates(sheet):
    for idx, row in enumerate(sheet.iter_rows(min_row=5, min_col=2, max_col=18, values_only=True), start=5):
        matricule = row[0]
        if matricule in sortie_matricules and matricule in mois_ref:
            mois_reference = ' | '.join([mois for mois in mois_ref[matricule]]).replace('|', ' | ')
            sheet.cell(row=idx, column=10, value=mois_reference)

# Mettre à jour les dates dans les feuilles TECHNIQUE et ADMINISTRATIF
update_dates(indemnite_technique_sheet)
update_dates(indemnite_administratif_sheet)

# R4 Enregistrer les modifications dans le fichier INDEMNITE ASTREINTE MOIS ANNEE
indemnite_workbook.save('INDEMNITE ASTREINTE SIEGE JUILLET 2025.xlsx')