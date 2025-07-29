import openpyxl
import sys
import os

def get_excel_filename(pattern):
    """Lire le nom du fichier Excel correspondant au pattern depuis excel_filenames.txt"""
    if os.path.exists('excel_filenames.txt'):
        with open('excel_filenames.txt', 'r') as f:
            filenames = [line.strip() for line in f.readlines()]
        
        # Chercher le fichier correspondant au pattern
        for filename in filenames:
            if pattern.lower() in filename.lower():
                return filename
    
    return None

# Tentative d'ouverture du fichier SORTIE.xlsx - DT
try:
    sortie_workbook = openpyxl.load_workbook('SORTIE.xlsx')
    sortie_sheet = sortie_workbook['HEURES SUPPLEMENTAIRES']
except FileNotFoundError:
    print("Erreur : Le fichier 'SORTIE.xlsx' est introuvable.")
    sys.exit(1)

# Chargement du fichier Excel - Récupération automatique du nom
etat_file_name = get_excel_filename('etat des heures supplementaires')
if not etat_file_name:
    print("Erreur : Impossible de trouver le fichier 'ETAT DES HEURES SUPPLEMENTAIRES'")
    sys.exit(1)

print(f"Tentative d'ouverture du fichier : {etat_file_name}")

try:
    etat_workbook = openpyxl.load_workbook(etat_file_name)
    etat_sheet = etat_workbook['HEURES SUPPLEMENTAIRES']
except FileNotFoundError:
    print(f"Erreur : Le fichier '{etat_file_name}' est introuvable.")
    sys.exit(1)

# Créer un dictionnaire pour stocker les mois de référence
mois_ref = {}

# Créer un ensemble pour stocker les matricules du fichier SORTIE.xlsx
sortie_matricules = set()

# Déterminer les numéros de colonnes en inspectant les en-têtes
headers = [cell.value for cell in sortie_sheet[1]]
matricule_col_idx = headers.index('Matricule') if 'Matricule' in headers else 0
mois_annee_col_idx = headers.index('Mois Année') if 'Mois Année' in headers else 1

# Parcourir les lignes de la feuille HEURES SUPPLEMENTAIRES de SORTIE.xlsx
for row in sortie_sheet.iter_rows(min_row=2, values_only=True):
    if len(row) > max(matricule_col_idx, mois_annee_col_idx):
        matricule = row[matricule_col_idx]
        mois_annee = row[mois_annee_col_idx]
        
        if matricule and mois_annee:
            sortie_matricules.add(matricule)
            if matricule in mois_ref:
                if mois_annee not in mois_ref[matricule]:
                    mois_ref[matricule].append(mois_annee)
            else:
                mois_ref[matricule] = [mois_annee]

# Trouver la colonne du matricule dans ETAT DES HEURES SUPPLEMENTAIRES
matricule_col = None
for row in range(1, 6):
    for col in range(1, 20):
        cell_value = etat_sheet.cell(row=row, column=col).value
        if cell_value and "matricule" in str(cell_value).lower():
            matricule_col = col
            break
    if matricule_col:
        break

if not matricule_col:
    print("Avertissement: En-tête 'Matricule' non trouvé, utilisation de la colonne C (3)")
    matricule_col = 3

# Parcourir les lignes à partir de la ligne 6
ligne_debut = 6
compteur_mises_a_jour = 0

for row in range(ligne_debut, etat_sheet.max_row + 1):
    matricule = etat_sheet.cell(row=row, column=matricule_col).value
    
    if matricule in sortie_matricules and matricule in mois_ref:
        mois_reference = " | ".join(mois_ref[matricule])
        etat_sheet.cell(row=row, column=14, value=mois_reference)
        compteur_mises_a_jour += 1

print(f'Fichier {etat_file_name} mis à jour avec succès.')

# Enregistrer les modifications dans le fichier
etat_workbook.save(etat_file_name)
print(f"Fichier {etat_file_name} mis à jour avec succès.")