
import openpyxl

# Fonction pour convertir le mois en un format de tri
def mois_en_numerique(mois):
    mois_numerique = {
        "JANVIER": 1,
        "FEVRIER": 2,
        "MARS": 3,
        "AVRIL": 4,
        "MAI": 5,
        "JUIN": 6,
        "JUILLET": 7,
        "AOUT": 8,
        "SEPTEMBRE": 9,
        "OCTOBRE": 10,
        "NOVEMBRE": 11,
        "DECEMBRE": 12
    }
    return mois_numerique.get(mois.upper(), 0)  # Return 0 if month not found

# Fonction pour extraire année et mois avec gestion d'erreurs
def process_date_string(date_str):
    parts = date_str.strip().split()
    if len(parts) >= 2:
        # If there are at least 2 parts, use the first two
        month = parts[0]
        year = parts[1]
        return (year, mois_en_numerique(month))
    elif len(parts) == 1:
        # If there's only one part, assume it's just a year
        return (parts[0], 0)
    else:
        # If there are no parts, return default values
        return ("0000", 0)

try:
    # R2 Charger le fichier Excel - DT
    wb = openpyxl.load_workbook('SORTIE.xlsx')
    
    # Parcourir chaque feuille
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # 1. First, let's swap the column headers
        temp_col2_header = sheet.cell(row=1, column=2).value
        sheet.cell(row=1, column=2).value = sheet.cell(row=1, column=3).value
        sheet.cell(row=1, column=3).value = temp_col2_header
        
        # 2. Swap the data in columns 2 and 3 for all rows
        for row_num in range(2, sheet.max_row + 1):
            # Temporarily store the value from column 2
            temp_col2_value = sheet.cell(row=row_num, column=2).value
            
            # Copy column 3 value to column 2
            sheet.cell(row=row_num, column=2).value = sheet.cell(row=row_num, column=3).value
            
            # Set column 3 value to the stored column 2 value
            sheet.cell(row=row_num, column=3).value = temp_col2_value
        
        # 3. Sort the month values in column 2 (now "Mois Année")
        for row_num in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_num, column=2).value
            
            if cell_value and isinstance(cell_value, str):
                try:
                    # Check different delimiters that might be present
                    if ' | ' in cell_value:
                        mois_list = cell_value.split(' | ')
                    elif '|' in cell_value:
                        mois_list = cell_value.split('|')
                    else:
                        # No delimiter found, treat as single value
                        mois_list = [cell_value]
                    
                    # Filtrer les chaînes vides et trier
                    mois_list = [x.strip() for x in mois_list if x.strip()]
                    
                    if mois_list:
                        # Sort by year (ascending) then by month (ascending)
                        mois_list.sort(key=process_date_string)
                        
                        # Join with proper format
                        sheet.cell(row=row_num, column=2).value = ' | '.join(mois_list)
                        
                except Exception as e:
                    print(f"Error processing row {row_num}: {e}")
                    continue  # Skip to the next row

    # R2 Enregistrer les modifications - DT
    wb.save('SORTIE.xlsx')
    print("SORTIE.xlsx réorganisé avec succès. Colonnes 2 et 3 interverties et mois triés.")
    
except Exception as e:
    print(f"Error in REORG-SORTIE.py: {e}")
